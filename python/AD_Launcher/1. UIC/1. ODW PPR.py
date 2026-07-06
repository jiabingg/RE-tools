#Help: Periodic Project Review using ODW data
import os
import oracledb
import tkinter as tk
from tkinter import messagebox, scrolledtext, filedialog
from tkinter import ttk
import tkinter.font
import ttkbootstrap as tb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta

# --- CONNECTION FIX: Enable Thick Mode ---
try:
    oracledb.init_oracle_client()
except Exception as e:
    print(f"Warning: Could not initialize Oracle thick client. {e}")
# -----------------------------------------


class OracleConnectionManager:
    def __init__(self):
        self._connections = {
            "odw": {
                "user": os.getenv("DB_USER_ODW", "rptguser"),
                "password": os.getenv("DB_PASSWORD_ODW", "allusers"),
                "dsn": "odw"
            },
            "sandbox": {
                "user": os.getenv("DB_USER_SANDBOX", "engsb"),
                "password": os.getenv("DB_PASSWORD_SANDBOX", "Engine33r_SB"),
                "dsn": "odw"
            },
            "openwells": {
                "user": os.getenv("DB_USER_OW", "gen_user"),
                "password": os.getenv("DB_PASSWORD_OW", "allusers"),
                "dsn": "owdb1"
            }
        }

    def get_connection(self, name):
        if name not in self._connections:
            raise ValueError(f"Unknown DB connection name: {name}")
        config = self._connections[name]
        try:
            return oracledb.connect(
                user=config['user'],
                password=config['password'],
                dsn=config['dsn']
            )
        except oracledb.Error as e:
            error_obj, = e.args
            raise ConnectionError(
                f"Failed to connect to Oracle DB '{name}': {error_obj.message}"
            ) from e


def format_well_api_list(raw_api_list):
    """Convert list of user-supplied APIs to a safe SQL IN list string."""
    cleaned = []
    for item in raw_api_list:
        if item is None:
            continue
        api = str(item).strip()
        if not api:
            continue
        if api not in cleaned:
            cleaned.append(api)
    if not cleaned:
        return None
    escaped = [f"'{x.replace(chr(39), chr(39)+chr(39))}'" for x in cleaned]
    return ", ".join(escaped)


def _is_api_column(col_name):
    """Return True if a column name looks like an API number column."""
    return 'API' in col_name.upper()


# --------------------------------------------------------------------------
# Mixin: Treeview display, clipboard copy (API-safe), and Excel export
# --------------------------------------------------------------------------
class TreeviewMixin:

    def display_results(self, df, apply_global_sort=True):
        for i in self.result_tree.get_children():
            self.result_tree.delete(i)

        if df.empty:
            self.result_tree["columns"] = []
            return

        if apply_global_sort:
            sort_cols = ['PRIM_PURP_TYPE_CDE', 'WELL_API_NBR']
            avail = [c for c in sort_cols if c in df.columns]
            if avail:
                try:
                    if 'WELL_API_NBR' in df.columns:
                        df['WELL_API_NBR'] = df['WELL_API_NBR'].astype(str)
                    df.sort_values(by=avail, inplace=True)
                except Exception:
                    pass

        # Force every API column to string so leading zeros survive
        for col in df.columns:
            if _is_api_column(col):
                df[col] = df[col].astype(str)

        columns = list(df.columns)
        self.result_tree["columns"] = columns
        self.result_tree["displaycolumns"] = columns

        try:
            fn = ttk.Style().lookup("Treeview", "font")
            tree_font = tkinter.font.Font(font=fn)
        except Exception:
            tree_font = tkinter.font.Font(family="TkDefaultFont", size=10)

        for col in columns:
            self.result_tree.heading(col, text=col, anchor="w")
            self.result_tree.column(col, width=tree_font.measure(col) + 20,
                                    stretch=False)

        for _, row in df.iterrows():
            vals = []
            for item in row:
                if isinstance(item, pd.Timestamp):
                    vals.append(item.strftime('%Y-%m-%d') if not pd.isna(item) else '')
                elif pd.isna(item):
                    vals.append('')
                else:
                    vals.append(item)
            self.result_tree.insert("", "end", values=vals)
            for i, item in enumerate(vals):
                w = tree_font.measure(str(item)) + 10
                cid = columns[i]
                if self.result_tree.column(cid, width=None) < w:
                    self.result_tree.column(cid, width=w)

    def clear_results(self):
        for i in self.result_tree.get_children():
            self.result_tree.delete(i)
        self.result_tree["columns"] = []
        self.current_data = None

    # --- Clipboard: wrap API columns so Excel keeps leading zeros ---
    def copy_to_clipboard(self):
        if self.current_data is None or self.current_data.empty:
            messagebox.showwarning("No Data", "No results to copy to clipboard.")
            return
        try:
            df_copy = self.current_data.copy()
            for col in df_copy.columns:
                if _is_api_column(col):
                    df_copy[col] = df_copy[col].apply(
                        lambda v: f'="{v}"' if pd.notna(v) and str(v).strip() else ''
                    )
            df_copy.to_clipboard(excel=True, index=False, header=True)
            messagebox.showinfo("Copy Success",
                                "Copied to clipboard (Excel format). "
                                "API columns are preserved as text.")
        except Exception as e:
            messagebox.showerror("Copy Error", f"Failed to copy: {e}")





# --------------------------------------------------------------------------
# Helper: build treeview + scrollbars
# --------------------------------------------------------------------------
def build_treeview(parent):
    tree_frame = tb.Frame(parent)
    sy = tb.Scrollbar(tree_frame, orient="vertical")
    sy.pack(side="right", fill="y")
    sx = tb.Scrollbar(tree_frame, orient="horizontal")
    sx.pack(side="bottom", fill="x")
    tree = ttk.Treeview(tree_frame, show="headings",
                         yscrollcommand=sy.set, xscrollcommand=sx.set)
    tree.pack(fill="both", expand=True)
    sy.config(command=tree.yview)
    sx.config(command=tree.xview)
    return tree_frame, tree


# --------------------------------------------------------------------------
# Helper: write a DataFrame to an openpyxl worksheet (API-safe)
# --------------------------------------------------------------------------
def _write_df_to_sheet(ws, df):
    """Write a DataFrame to an openpyxl worksheet with formatting."""
    columns = list(df.columns)
    api_col_indices = {i for i, c in enumerate(columns) if _is_api_column(c)}

    header_font = Font(name="Arial", bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    data_font = Font(name="Arial", size=10)

    for c_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, (col_name, val) in enumerate(zip(columns, row), start=1):
            if isinstance(val, pd.Timestamp):
                cell_val = val.strftime('%Y-%m-%d') if not pd.isna(val) else None
            elif pd.isna(val):
                cell_val = None
            else:
                cell_val = val

            cell = ws.cell(row=r_idx, column=c_idx, value=cell_val)
            cell.font = data_font

            if (c_idx - 1) in api_col_indices:
                cell.number_format = '@'
                cell.value = str(cell_val) if cell_val is not None else ''

    for c_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for r_idx in range(2, min(ws.max_row + 1, 200)):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"


# --------------------------------------------------------------------------
# Export ALL tabs into one xlsx workbook (one sheet per tab)
# --------------------------------------------------------------------------
def export_all_tabs(app):
    """Collect data from every data tab and write to a single .xlsx file."""
    tabs = [
        ("Basic Data",       app.basic_tab),
        ("Top Perf",         app.perf_tab),
        ("Summary",          app.summary_tab),
        ("Avg Tubing Pres",  app.tubing_tab),
        ("Monthly Prod Inj", app.prod_inj_tab),
        ("Daily Inj Pres",   app.daily_tab),
    ]

    # Check that at least one tab has data
    has_data = any(
        getattr(t, 'current_data', None) is not None
        and not t.current_data.empty
        for _, t in tabs
    )
    if not has_data:
        messagebox.showwarning("No Data",
                               "No data loaded on any tab. "
                               "Click 'Load All Data' first.")
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Export All Tabs to Spreadsheet"
    )
    if not file_path:
        return

    try:
        wb = Workbook()
        # Remove the default empty sheet that openpyxl creates
        wb.remove(wb.active)

        sheets_written = 0
        for sheet_name, tab in tabs:
            df = getattr(tab, 'current_data', None)
            if df is None or df.empty:
                continue
            ws = wb.create_sheet(title=sheet_name)
            _write_df_to_sheet(ws, df.copy())
            sheets_written += 1

        if sheets_written == 0:
            messagebox.showwarning("No Data", "All tabs are empty — nothing to export.")
            return

        wb.save(file_path)
        messagebox.showinfo("Export Success",
                            f"Exported {sheets_written} tab(s) to:\n{file_path}")

    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export: {e}")


# --------------------------------------------------------------------------
# Helper: bottom button bar (Copy + Export All) — shared by all data tabs
# --------------------------------------------------------------------------
def build_button_bar(parent, tab):
    bar = tb.Frame(parent)
    tb.Button(bar, text="Copy to Clipboard", command=tab.copy_to_clipboard,
              bootstyle="secondary").pack(side="left", padx=5)
    tb.Button(bar, text="Export All Tabs to Spreadsheet",
              command=lambda: export_all_tabs(tab.app),
              bootstyle="success").pack(side="left", padx=5)
    return bar


# =========================================================================
#  TAB 1 – Well API Input
# =========================================================================
class WellAPITab(tb.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._build()

    def _build(self):
        tb.Label(self,
                 text="Enter Well APIs (one per line — all tabs load automatically):",
                 font=("Helvetica", 14)).pack(pady=10)

        self.well_api_text = scrolledtext.ScrolledText(
            self, wrap=tk.WORD, width=50, height=25, font=("Courier New", 10))
        self.well_api_text.pack(pady=10)

        default_apis = ["0401920171", "0401922081", "0401922236"]
        self.well_api_text.insert(tk.END, "\n".join(default_apis))

        tb.Button(self, text="Load All Data", command=self._load_all,
                  bootstyle="primary").pack(pady=10)

        self.status_label = tb.Label(self, text="", font=("Helvetica", 11))
        self.status_label.pack(pady=5)

    def get_apis(self):
        raw = self.well_api_text.get("1.0", tk.END).strip().split('\n')
        apis = list(dict.fromkeys(a.strip() for a in raw if a.strip()))
        self.app.shared_data["well_apis"] = apis
        return apis

    def _load_all(self):
        apis = self.get_apis()
        if not apis:
            messagebox.showwarning("Input Error",
                                   "Please enter at least one Well API number.")
            return

        self.status_label.config(text=f"Loading data for {len(apis)} API(s)…")
        self.update_idletasks()

        tabs_to_load = [
            ("Basic Data",        self.app.basic_tab),
            ("Top Perf",          self.app.perf_tab),
            ("Summary",           self.app.summary_tab),
            ("Avg Tubing Pres",   self.app.tubing_tab),
            ("Monthly Prod/Inj",  self.app.prod_inj_tab),
            ("Daily Inj/Pres",    self.app.daily_tab),
        ]

        errors = []
        for name, tab in tabs_to_load:
            try:
                self.status_label.config(text=f"Loading {name}…")
                self.update_idletasks()
                tab.pull_data()
            except Exception as e:
                errors.append(f"{name}: {e}")

        if errors:
            self.status_label.config(text="Done (some tabs had errors)")
            messagebox.showwarning("Load Warnings",
                                   "Errors on these tabs:\n\n" + "\n".join(errors))
        else:
            self.status_label.config(text="All tabs loaded successfully.")


# =========================================================================
#  Shared query executor
# =========================================================================
class _QueryTab:
    """Mixin that adds a shared _execute helper."""
    conn_manager = OracleConnectionManager()

    def _execute(self, sql, date_cols=None, sort=True):
        try:
            conn = self.conn_manager.get_connection('odw')
            cursor = conn.cursor()
            cursor.execute(sql)
            if cursor.description:
                rows = cursor.fetchall()
                columns = [col[0] for col in cursor.description]
                df = pd.DataFrame(rows, columns=columns)
                for col in (date_cols or []):
                    if col in df.columns:
                        df[col] = pd.to_datetime(df[col], errors='coerce')
                self.display_results(df, apply_global_sort=sort)
                self.current_data = df
            else:
                self.clear_results()
            cursor.close()
            conn.close()
        except ConnectionError as e:
            messagebox.showerror("Connection Error", str(e))
            self.clear_results()
        except oracledb.Error as e:
            error_obj, = e.args
            messagebox.showerror("Database Error",
                                 f"Oracle Error: {error_obj.message}")
            self.clear_results()
        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error: {e}")
            self.clear_results()


# =========================================================================
#  TAB 2 – Well Basic Data
# =========================================================================
class WellBasicDataTab(tb.Frame, TreeviewMixin, _QueryTab):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.current_data = None
        self._build()

    def _build(self):
        tb.Label(self, text="Well Basic Data",
                 font=("Helvetica", 16, "bold")).pack(pady=10)
        self.tree_frame, self.result_tree = build_treeview(self)
        self.tree_frame.pack(pady=10, padx=20, fill="both", expand=True)
        build_button_bar(self, self).pack(pady=10)

    def pull_data(self):
        formatted = format_well_api_list(self.app.api_tab.get_apis())
        if not formatted:
            self.clear_results(); return
        sql = f"""
SELECT
    wd.wlbr_nme                    AS well_name,
    cd.opnl_fld                    AS field_name,
    cd.cmpl_nme                    AS completion_name,
    cd.well_api_nbr                AS api_number,
    wd.wlbr_api_suff_nbr           AS wellbore_suffix,
    wd.wlbr_incl_type_desc         AS wellbore_type,
    cd.prim_purp_type_cde          AS well_type,
    cd.cmpl_state_type_cde         AS status,
    cd.in_svc_indc                 AS in_service,
    cd.init_prod_dte               AS initial_prod_date
FROM dwrptg.cmpl_dmn cd
JOIN dwrptg.wlbr_dmn wd ON cd.well_fac_id = wd.well_fac_id
WHERE cd.actv_indc = 'Y' AND cd.well_api_nbr IN ({formatted})
ORDER BY cd.well_api_nbr, wd.wlbr_api_suff_nbr, cd.cmpl_nme
"""
        self._execute(sql, date_cols=['INITIAL_PROD_DATE'])


# =========================================================================
#  TAB 3 – Top Perf Data
# =========================================================================
class TopPerfTab(tb.Frame, TreeviewMixin, _QueryTab):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.current_data = None
        self._build()

    def _build(self):
        tb.Label(self, text="Top Perf Query",
                 font=("Helvetica", 16, "bold")).pack(pady=10)
        self.tree_frame, self.result_tree = build_treeview(self)
        self.tree_frame.pack(pady=10, padx=20, fill="both", expand=True)
        build_button_bar(self, self).pack(pady=10)

    def pull_data(self):
        formatted = format_well_api_list(self.app.api_tab.get_apis())
        if not formatted:
            self.clear_results(); return
        sql = f"""
WITH T AS (
    SELECT cd.cmpl_nme, cd.cmpl_fac_id, cd.well_fac_id,
           wd.well_api_nbr, cd.engr_strg_nme
    FROM dwrptg.cmpl_dmn cd
    JOIN dwrptg.well_dmn wd ON cd.well_fac_id = wd.well_fac_id
    WHERE cd.actv_indc = 'Y'
      AND wd.actv_indc = 'Y'
      AND wd.well_api_nbr IN ({formatted})
      AND cd.cmpl_state_type_cde IN ('OPNL', 'TA', 'ABND')
),
perfs AS (
    SELECT t.well_api_nbr, t.cmpl_nme, t.cmpl_fac_id,
           t.well_fac_id, t.engr_strg_nme,
           MIN(opg.top_md_qty) AS top_perf,
           MAX(opg.btm_md_qty) AS btm_perf
    FROM T
    JOIN dwrptg.wlbr_dmn wd ON t.well_fac_id = wd.well_fac_id
    JOIN dwrptg.actl_wlbr_opg_ntvl_dmn opg ON wd.wlbr_fac_id = opg.wlbr_fac_id
    GROUP BY t.well_api_nbr, t.cmpl_nme, t.cmpl_fac_id,
             t.well_fac_id, t.engr_strg_nme
),
surveys AS (
    SELECT wd.well_fac_id, d.md_qty AS svy_md, d.tvd_qty AS svy_tvd
    FROM dwrptg.dsvy_pt_dmn d
    JOIN dwrptg.wlbr_dmn wd ON d.wlbr_fac_id = wd.wlbr_fac_id
    WHERE wd.well_fac_id IN (SELECT well_fac_id FROM T)
      AND d.tvd_qty IS NOT NULL AND d.md_qty IS NOT NULL
      AND d.tvd_qty <= d.md_qty
),
top_above AS (
    SELECT p.cmpl_fac_id, s.svy_md, s.svy_tvd,
           ROW_NUMBER() OVER (PARTITION BY p.cmpl_fac_id ORDER BY s.svy_md DESC) AS rn
    FROM perfs p JOIN surveys s ON s.well_fac_id = p.well_fac_id
    WHERE s.svy_md <= p.top_perf
),
top_below AS (
    SELECT p.cmpl_fac_id, s.svy_md, s.svy_tvd,
           ROW_NUMBER() OVER (PARTITION BY p.cmpl_fac_id ORDER BY s.svy_md ASC) AS rn
    FROM perfs p JOIN surveys s ON s.well_fac_id = p.well_fac_id
    WHERE s.svy_md > p.top_perf
),
btm_above AS (
    SELECT p.cmpl_fac_id, s.svy_md, s.svy_tvd,
           ROW_NUMBER() OVER (PARTITION BY p.cmpl_fac_id ORDER BY s.svy_md DESC) AS rn
    FROM perfs p JOIN surveys s ON s.well_fac_id = p.well_fac_id
    WHERE s.svy_md <= p.btm_perf
),
btm_below AS (
    SELECT p.cmpl_fac_id, s.svy_md, s.svy_tvd,
           ROW_NUMBER() OVER (PARTITION BY p.cmpl_fac_id ORDER BY s.svy_md ASC) AS rn
    FROM perfs p JOIN surveys s ON s.well_fac_id = p.well_fac_id
    WHERE s.svy_md > p.btm_perf
)
SELECT p.well_api_nbr, p.cmpl_nme, p.engr_strg_nme,
       p.top_perf,
       LEAST(p.top_perf,
             ROUND(CASE
                 WHEN ta.svy_md IS NOT NULL AND tb.svy_md IS NOT NULL
                      AND tb.svy_md != ta.svy_md
                 THEN ta.svy_tvd + (p.top_perf - ta.svy_md)
                      * (tb.svy_tvd - ta.svy_tvd) / (tb.svy_md - ta.svy_md)
                 WHEN ta.svy_md IS NOT NULL THEN ta.svy_tvd
                 ELSE p.top_perf END, 1)) AS top_perf_tvd,
       p.btm_perf,
       LEAST(p.btm_perf,
             ROUND(CASE
                 WHEN ba.svy_md IS NOT NULL AND bb.svy_md IS NOT NULL
                      AND bb.svy_md != ba.svy_md
                 THEN ba.svy_tvd + (p.btm_perf - ba.svy_md)
                      * (bb.svy_tvd - ba.svy_tvd) / (bb.svy_md - ba.svy_md)
                 WHEN ba.svy_md IS NOT NULL THEN ba.svy_tvd
                 ELSE p.btm_perf END, 1)) AS btm_perf_tvd
FROM perfs p
LEFT JOIN top_above ta ON p.cmpl_fac_id = ta.cmpl_fac_id AND ta.rn = 1
LEFT JOIN top_below tb ON p.cmpl_fac_id = tb.cmpl_fac_id AND tb.rn = 1
LEFT JOIN btm_above ba ON p.cmpl_fac_id = ba.cmpl_fac_id AND ba.rn = 1
LEFT JOIN btm_below bb ON p.cmpl_fac_id = bb.cmpl_fac_id AND bb.rn = 1
ORDER BY p.cmpl_nme
"""
        self._execute(sql)


# =========================================================================
#  TAB 4 – Performance Summary  (simplified — data table only)
# =========================================================================
class PerformanceSummaryTab(tb.Frame, TreeviewMixin, _QueryTab):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.current_data = None
        self._build()

    def _build(self):
        tb.Label(self, text="Performance Summary",
                 font=("Helvetica", 16, "bold")).pack(pady=10)
        self.tree_frame, self.result_tree = build_treeview(self)
        self.tree_frame.pack(pady=10, padx=20, fill="both", expand=True)
        build_button_bar(self, self).pack(pady=10)

    def pull_data(self):
        formatted = format_well_api_list(self.app.api_tab.get_apis())
        if not formatted:
            self.clear_results(); return
        sql = f"""
WITH T1 AS (
    SELECT cmpl_fac_id, eftv_dttm AS last_inj_dte FROM (
        SELECT cmpl_fac_id, eftv_dttm,
               DENSE_RANK() OVER (PARTITION BY cmpl_fac_id ORDER BY eftv_dttm DESC) AS rnk
        FROM cmpl_mnly_fact
        WHERE aloc_wtr_inj_dly_rte_qty > 0 OR aloc_stm_inj_dly_rte_qty > 0
    ) WHERE rnk = 1
),
T2 AS (
    SELECT cmpl_fac_id, eftv_dttm AS last_prod_dte FROM (
        SELECT cmpl_fac_id, eftv_dttm,
               DENSE_RANK() OVER (PARTITION BY cmpl_fac_id ORDER BY eftv_dttm DESC) AS rnk
        FROM cmpl_mnly_fact
        WHERE aloc_gros_prod_dly_rte_qty > 0
    ) WHERE rnk = 1
)
SELECT wd.well_nme, wd.well_api_nbr, wd.fld_nme,
       cd.init_prod_dte, cd.init_inj_dte, cd.prim_purp_type_cde,
       cd.ENGR_STRG_NME,
       t1.last_inj_dte, t2.last_prod_dte,
       cd.CMPL_STATE_TYPE_DESC, cd.CMPL_STATE_EFTV_DTTM
FROM well_dmn wd
JOIN cmpl_dmn cd ON wd.well_fac_id = cd.well_fac_id
LEFT JOIN cmpl_non_ver_dmn cnd ON cd.cmpl_fac_id = cnd.cmpl_fac_id
LEFT JOIN curr_cmpl_opnl_stat os ON cd.cmpl_fac_id = os.cmpl_fac_id
LEFT JOIN T1 ON cd.cmpl_fac_id = T1.cmpl_fac_id
LEFT JOIN T2 ON cd.cmpl_fac_id = T2.cmpl_fac_id
WHERE cd.actv_indc = 'Y' AND wd.actv_indc = 'Y'
  AND wd.well_api_nbr IN ({formatted})
  AND cd.prim_purp_type_cde IN ('PROD', 'INJ')
"""
        self._execute(sql, date_cols=['LAST_INJ_DTE', 'LAST_PROD_DTE',
                                       'INIT_INJ_DTE', 'INIT_PROD_DTE',
                                       'CMPL_STATE_EFTV_DTTM'])


# =========================================================================
#  TAB 5 – Avg Tubing Pressure & Inj Vol
# =========================================================================
class TubingPressureTab(tb.Frame, TreeviewMixin, _QueryTab):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.current_data = None
        self._build()

    def _build(self):
        tb.Label(self, text="Avg Injection Volume and Wlhd Tubing Pressure",
                 font=("Helvetica", 16, "bold")).pack(pady=10)

        # Average pressure readout
        avg_frame = tb.Frame(self)
        avg_frame.pack(pady=5, padx=20, fill="x")
        tb.Label(avg_frame, text="Overall Avg Tubing Pressure:").pack(side="left", padx=5)
        self.avg_pressure_label = tb.Label(avg_frame, text="N/A",
                                            bootstyle="success",
                                            font=("TkDefaultFont", 10, "bold"))
        self.avg_pressure_label.pack(side="left", padx=5)

        self.tree_frame, self.result_tree = build_treeview(self)
        self.tree_frame.pack(pady=10, padx=20, fill="both", expand=True)
        build_button_bar(self, self).pack(pady=10)

    def pull_data(self):
        formatted = format_well_api_list(self.app.api_tab.get_apis())
        if not formatted:
            self.clear_results(); self.avg_pressure_label.config(text="N/A"); return
        sql = f"""
SELECT wd.well_nme, wd.well_api_nbr, cd.cmpl_nme, cd.cmpl_fac_id,
    AVG(CASE WHEN cf.aloc_stm_inj_vol_qty > 0
         THEN cf.aloc_stm_inj_vol_qty END) AS avg_stm_inj_vol,
    ROUND(AVG(CASE WHEN cf.aloc_wtr_inj_vol_qty > 0
              THEN cf.aloc_wtr_inj_vol_qty END), 2) AS avg_wtr_inj_vol,
    ROUND(AVG(CASE WHEN cf.wlhd_tbg_prsr_qty > 0
              THEN cf.wlhd_tbg_prsr_qty END), 2) AS avg_wlhd_tbg_prsr
FROM well_dmn wd
JOIN cmpl_dmn cd ON wd.well_fac_id = cd.well_fac_id
JOIN cmpl_dly_fact cf ON cd.cmpl_fac_id = cf.cmpl_fac_id
WHERE wd.actv_indc = 'Y' AND cd.actv_indc = 'Y'
    AND cf.eftv_dttm >= TRUNC(SYSDATE) - 60
    AND wd.well_api_nbr IN ({formatted})
GROUP BY wd.well_nme, wd.well_api_nbr, cd.cmpl_nme, cd.cmpl_fac_id
"""
        self._execute(sql)
        # Calculate overall average after data loads
        if self.current_data is not None and not self.current_data.empty:
            col = next((c for c in self.current_data.columns
                        if c.upper() == 'AVG_WLHD_TBG_PRSR'), None)
            if col:
                vals = pd.to_numeric(self.current_data[col], errors='coerce').dropna()
                if not vals.empty:
                    self.avg_pressure_label.config(text=f"{vals.mean():.1f}")
                    return
        self.avg_pressure_label.config(text="N/A")


# =========================================================================
#  TAB 6 – Monthly Production & Injection
# =========================================================================
class ProductionInjectionTab(tb.Frame, TreeviewMixin, _QueryTab):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.current_data = None
        self._build()

    def _build(self):
        tb.Label(self, text="Production and Injection Data",
                 font=("Helvetica", 16, "bold")).pack(pady=10)
        self.tree_frame, self.result_tree = build_treeview(self)
        self.tree_frame.pack(pady=10, padx=20, fill="both", expand=True)
        build_button_bar(self, self).pack(pady=10)

    def pull_data(self):
        formatted = format_well_api_list(self.app.api_tab.get_apis())
        if not formatted:
            self.clear_results(); return
        sql = f"""
SELECT
    wd.well_nme AS "WELL NAME",
    wd.well_api_nbr AS "WELL API",
    cf.eftv_dttm AS "DATE",
    cf.aloc_oil_prod_dly_rte_qty AS "OIL PROD BOPD",
    cf.aloc_wtr_prod_dly_rte_qty AS "WATER PROD BWPD",
    cf.aloc_gas_prod_dly_rte_qty AS "GAS PROD MCFD",
    cf.aloc_stm_inj_dly_rte_qty AS "STEAM INJ Per Day",
    cf.aloc_wtr_inj_dly_rte_qty AS "WATER INJ Per Day",
    cf.aloc_gas_inj_dly_rte_qty AS "GAS INJ Per Day"
FROM well_dmn wd
JOIN cmpl_dmn cd ON wd.well_fac_id = cd.well_fac_id
JOIN cmpl_mnly_fact cf ON cd.cmpl_fac_id = cf.cmpl_fac_id
WHERE cd.actv_indc = 'Y' AND wd.actv_indc = 'Y'
    AND wd.well_api_nbr IN ({formatted})
    AND cf.eftv_dttm >= ADD_MONTHS(TRUNC(SYSDATE), -62)
    AND cf.eftv_dttm <= TRUNC(SYSDATE)
ORDER BY wd.well_api_nbr, cf.eftv_dttm
"""
        self._execute(sql, date_cols=['DATE'], sort=False)


# =========================================================================
#  TAB 7 – Daily Injection & Tubing Pressure
# =========================================================================
class DailyInjectionPressureTab(tb.Frame, TreeviewMixin, _QueryTab):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.current_data = None
        self._build()

    def _build(self):
        tb.Label(self, text="Daily Injection and Tubing Pressure (Full Data)",
                 font=("Helvetica", 16, "bold")).pack(pady=10)
        self.tree_frame, self.result_tree = build_treeview(self)
        self.tree_frame.pack(pady=10, padx=20, fill="both", expand=True)
        build_button_bar(self, self).pack(pady=10)

    def pull_data(self):
        formatted = format_well_api_list(self.app.api_tab.get_apis())
        if not formatted:
            self.clear_results(); return
        sql = f"""
SELECT wd.well_nme, wd.well_api_nbr, cd.cmpl_nme, cd.cmpl_fac_id,
    cf.eftv_dttm,
    cf.aloc_stm_inj_vol_qty,
    cf.aloc_wtr_inj_vol_qty,
    cf.wlhd_tbg_prsr_qty
FROM well_dmn wd
JOIN cmpl_dmn cd ON wd.well_fac_id = cd.well_fac_id
JOIN cmpl_dly_fact cf ON cd.cmpl_fac_id = cf.cmpl_fac_id
WHERE wd.actv_indc = 'Y' AND cd.actv_indc = 'Y'
    AND cf.eftv_dttm >= TRUNC(SYSDATE) - 60
    AND wd.well_api_nbr IN ({formatted})
ORDER BY cf.eftv_dttm
"""
        self._execute(sql, date_cols=['EFTV_DTTM'], sort=False)


# =========================================================================
#  MAIN APPLICATION
# =========================================================================
class MainApplication(tb.Window):
    def __init__(self):
        super().__init__(themename="flatly")
        self.title("Periodic Project Review")
        self.geometry("1600x1200")

        s = ttk.Style()
        s.configure("TButton", font=("Helvetica", 12, "bold"))
        s.configure("TNotebook.Tab", font=("Helvetica", 11, "bold"),
                    padding=[12, 6])

        self.shared_data = {}

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.api_tab = WellAPITab(self.notebook, self)
        self.notebook.add(self.api_tab, text="  Well APIs  ")

        self.basic_tab = WellBasicDataTab(self.notebook, self)
        self.notebook.add(self.basic_tab, text="  Basic Data  ")

        self.perf_tab = TopPerfTab(self.notebook, self)
        self.notebook.add(self.perf_tab, text="  Top Perf  ")

        self.summary_tab = PerformanceSummaryTab(self.notebook, self)
        self.notebook.add(self.summary_tab, text="  Summary  ")

        self.tubing_tab = TubingPressureTab(self.notebook, self)
        self.notebook.add(self.tubing_tab, text="  Avg Tubing Pres  ")

        self.prod_inj_tab = ProductionInjectionTab(self.notebook, self)
        self.notebook.add(self.prod_inj_tab, text="  Monthly Prod/Inj  ")

        self.daily_tab = DailyInjectionPressureTab(self.notebook, self)
        self.notebook.add(self.daily_tab, text="  Daily Inj/Pres  ")


if __name__ == "__main__":
    app = MainApplication()
    app.mainloop()